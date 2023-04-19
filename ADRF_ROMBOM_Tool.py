import streamlit as st
import os
import pandas as pd
#import numpy as np
#import pdb
import math
#import openpyxl
import openpyxl
#using open py xl to load it in
from openpyxl import load_workbook
#importing open py xl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import glob
#from natsort import natsorted
import pytz
from pytz import timezone

from copy import copy
import streamlit
from io import BytesIO
from tempfile import NamedTemporaryFile
from pathlib import Path
import datetime

import gspread
from oauth2client.service_account import ServiceAccountCredentials

import smtplib
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate
from email import encoders
from email.message import EmailMessage




# The way this tool works is that it creates a new worksheet using openpyxl, it also reads from a template file that should be in the same directory
# The primary function that this tool automates is that it fills in the rows and formating from the template file
# everything else is hardcoded into the actual rows
# evenutally the goal is to reference any template sheet and organize everything into lists, which will then auto populate which should allow for increased flexibility on the input file
# potentially even uploading one to the tool and having it read it and then eventually a google sheets api call so the system can be adjusted as needed

	
ver_num = "Alpha 1.3"
#email configuration
daset_email = "websupport@adrftech.com"
daset_password = "wjdzdglbmqqbkhgk"

send_to_email_addresses = ["daset@adrftech.com","sales@adrftech.com"]

st.set_page_config(page_title="ADRF ROM BOM Tool",layout="wide")
relevent_col = ['A','B','C','D','E']
antenna_mount = 'Non-Pentrating'
#sectors = 1
donor_low = 0
donor_high = 0
seven_h_mhz_count = 0
seven_h_l_mhz_count = 0
seven_h_u_mhz_count = 0
cellular_count = 0
smr_count = 0
pcs_count = 0
aws_count = 0
brs_count = 0
att_bands = []
tmob_bands = []
vzw_bands = []
vhf_duplexing = "No"
uhf_duplexing = "No"
siso_or_mimo ="SISO"
mimo_streams = 1
north_east_region = ['CT', 'DC', 'DE', 'MA', 'MD', 'ME','NH', 'NJ', 'NY', 'PA', 'RI','VA', 'VT','WV', 'Canada-North']
south_east_region = ['AL', 'FL', 'GA','LA','MS', 'NC', 'SC', 'TN', 'Latin-America']
central_region = ['AR', 'Canada-Central','IA', 'IL', 'IN', 'KS', 'KY', 'MI', 'MN', 'MO','ND', 'NE', 'OH', 'OK', 'SD', 'TX','WI']
west_coast_region = ['AK', 'AZ', 'CA', 'CO', 'HI', 'ID','MT', 'NM', 'NV', 'OR', 'UT', 'WA', 'WY', 'Canada-West']
rsm_email=""
project_region=""

#Change these Emails for new RSMs
north_east_rsm = "lpeisel@adrftech.com"
south_east_rsm = "jlilienfeld@adrftech.com"
central_rsm = "troth@adrftech.com"
west_coast_rsm = "mjones@adrftech.com"

st.title('ADRF ROM BOM Tool')

st.write("Tutorial Video: https://www.youtube.com/watch?v=Yd-otkp2cO8")
# st.write('---')
# st.write("####")

with st.container():
	left_column, right_column = st.columns(2)
	with left_column:
		project_name = st.text_input('Project Name (Do not include character: /)', 'Project')
		commercial_or_ps = st.selectbox(
			'Is this for Commercial or Public Safety?',
			('Commercial', 'Public Safety'))

		square_feet = st.number_input('What is the approximate square footage of the project?',5000,step=1)
		project_state = st.selectbox(
	'What State is the project located in?',
	('AK', 'AL', 'AR', 'AZ', 'CA', 'CO', 'CT', 'DC', 'DE', 'FL', 'GA',
		 'HI', 'IA', 'ID', 'IL', 'IN', 'KS', 'KY', 'LA', 'MA', 'MD', 'ME',
		 'MI', 'MN', 'MO', 'MS', 'MT', 'NC', 'ND', 'NE', 'NH', 'NJ', 'NM',
		 'NV', 'NY', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX',
		 'UT', 'VA', 'VT', 'WA', 'WI', 'WV', 'WY','Canada-West','Canada-Central','Canada-North','Latin-America'))
		#num_of_buildings = st.number_input('How many buildings is the project?',1,step=1)
		#num_of_floors = st.number_input('How many floors is the building?',1,step=1)
		if project_state in north_east_region:
			st.write("Project Located in North East Region")
			project_region = "North East Region"
			rsm_email = north_east_rsm
		elif project_state in south_east_region:
			st.write("Project Located in South East Region")
			project_region = "South East Region"
			rsm_email = south_east_rsm
		elif project_state in central_region:
			st.write("Project Located in Central Region")
			project_region = "Central Region"
			rsm_email = central_rsm
		elif project_state in west_coast_region:
			st.write("Project Located in West Coast Region")
			project_region = "West Coast Region"
			rsm_email = west_coast_rsm
		send_to_email_addresses.append(rsm_email)
		
		
		
		


	#with left_middle_column:
		num_of_floors = 1

		num_of_buildings = st.number_input('How many buildings?',1,10,step=1)
		if num_of_buildings == 1:
			num_of_floors = st.number_input('How many floors is the building?',1,step=1)

		if num_of_buildings > 1:
			num_of_floors = 0
			all_floors = []
			with st.expander("Building List"):
				for i in range(num_of_buildings):
					
					all_floors.append(st.number_input(f'How many floors is building {i+1}?',1,step=1))
			for floors in all_floors:
				num_of_floors += floors





	#with right_middle_column:
	with right_column:
		
		bands = set([])
		if commercial_or_ps == "Commercial":
			carriers = st.multiselect(
				'Select Carriers',
				['AT&T','T-Mobile','Verizon'],
				['AT&T','T-Mobile','Verizon'])
			if 'AT&T' in carriers:
				att_bands = st.multiselect(
					'Select your AT&T freqency bands',
					['700 MHz', 'Cellular','PCS','AWS'],
					['700 MHz', 'Cellular','PCS','AWS'])
			if 'T-Mobile' in carriers:
				tmob_bands = st.multiselect(
					'Select your T-Mobile freqency bands',
					['700 MHz','SMR','PCS','AWS', 'BRS'],
					['700 MHz','PCS','AWS', 'BRS'])
			if 'Verizon' in carriers:
				vzw_bands = st.multiselect(
					'Select your Verizon freqency bands',
					['700 MHz', 'Cellular','PCS','AWS'],
					['700 MHz', 'Cellular','PCS','AWS'])
		
			if "700 MHz" in att_bands:
				seven_h_l_mhz_count += 1
			if "700 MHz" in tmob_bands:
				seven_h_l_mhz_count += 1
			if "700 MHz" in vzw_bands:
				seven_h_u_mhz_count += 1

			for carrier in [att_bands,tmob_bands,vzw_bands]:
				for freq in carrier:
					bands.add(freq)

					if freq == '700 MHz':
						seven_h_mhz_count+=1
						
					if freq == 'SMR':
						smr_count+=1
					if freq == 'Cellular':
						cellular_count+=1
					if freq == 'PCS':
						pcs_count+=1
					if freq == 'AWS':
						aws_count+=1
					if freq == 'BRS':
						brs_count+=1

		if commercial_or_ps == "Public Safety":
			ps_bands = st.multiselect(
			'Selcect your freqency bands',
			['VHF', 'UHF','700/800 MHz PS'],
			['700/800 MHz PS'])
			if 'VHF' in ps_bands:
				vhf_duplexing = st.radio("Is VHF Duplexed?",('Yes', 'No'))
			if 'UHF' in ps_bands:
				uhf_duplexing = st.radio("Is UHF Duplexed?",('Yes', 'No'))	
			backup_time = st.radio("Battery Backup Time: ",('12 Hr', '24 Hr'))

			antenna_mount = st.radio("Antenna Mounting Type:",('Non-Pentrating', 'Wall Mount'))
			if 'VHF' in ps_bands or 'UHF' in ps_bands:
				st.write("Validation Form Required: https://adrftech.com/psrvu/")
			epo_switch = st.checkbox(label="EPO Switch Required",value=False)
			wall_mount_battery = st.checkbox(label="Wall Mounted Battery Backup",value=False)



	#with right_column:
		if square_feet > 100000 or num_of_buildings > 1:
			if commercial_or_ps == "Commercial":
				signal_source = st.selectbox(
					'What is the Signal Source?',
					('Repeaters', 'Base Station/Small Cell/eFemto'))
				mpr_or_hpr  = st.selectbox(
					'Medium Power Remote or High Power Remote',
					('Medium Power Remote', 'High Power Remote'))
				sectors = st.number_input('Number of Sectors',1,step=1)
				
				#this can be updated to st.number input for more than 2 streams of MIMO

				siso_or_mimo = st.radio("SISO or MIMO?",('SISO', 'MIMO'))
				if siso_or_mimo == "MIMO":
					mimo_streams = 2
					signal_source = "Base Station/Small Cell/eFemto"


		with st.expander("Limitations"):
			st.write("""
				\nThis Bill of Materials is based on information (“Customer Information”) (e.g., rough order of magnitude (ROM) estimate, preliminary design, customer-provided equipment list) provided by the customer (“Customer”) and is intended for budgetary purposes only.  
				\nADRF does not represent or warrant the accuracy of the Customer Information.  
				\nThe Customer is solely responsible for the accuracy of the Customer Information provided to ADRF.  
				\nThe Customer agrees to hold ADRF harmless for any inaccurate Customer Information provided to ADRF.
				""")
		with st.expander("Please reach out to daset@adrftech.com if your project falls into the following: "):
			st.write("""
			\nLarger than 10 buildings
			\nLarger than 6.4M Square Feet
			\nC-Band is required
			\nSMR (862-869MHz) or 900MHz Required
			\nCandian Commercial Freqencies Required
			\nLatin America Freqencies Required
			""")
#		st.write(ver_num)
#		with st.expander("Change Logs"):
#			st.write("""
#
#				1.3\n
#				>DA:\n
#				>>Added MIMO option\n
#				>>changed signal source language\n
#				>>added emailing functionality\n
#				>>added wall mount battery backup options\n
#				>>added customer email
#				>>added canadian BRS freqencies for passive/MPR active
#
#
#				1.2\n
#				>DA:\n
#				>>Added google sheets support\n
#				>>fixed fire OEU not showing up\n
#				>>added EPO Switch option\n
#				>>Adjusted layout\n
#
#				1.1\n
#				>DA:\n
#				>>Updated to new 2023 Template\n
#				>>added logo and disclaimer \n
#				>>fixed donor coax/connectors not showing up \n
#				>>fixed issue with fiber patch panel\n
#				>>fixed issue with PS Battery backups/battery cables\n
#				>>fixed adapter plate count\n
#				>>fixed remote passives count\n
#				>>set date to central time\n
#				>>updated date foramtting \n
#				>>fixed ps 4.3 jumper count\n
#				>>updated fiber adapter plate count\n
#				>>removed couplers/splitters showing qty 0\n
#
#				1.0:\n
#				>DA:\n
#				>>Added Changelog\n
#				>>Fixed Fiber not showing up for PS\n
#				>>Fixed VHF/UHF passives showing up for 800 only\n
#				>>Added Donor connectors to ancillaries\n
#
#				
#			""")
#
		




relevent_col=['A','B','C','D','E']

#Importing template file

template = load_workbook(filename='ADRF_Template_ROM BOM_2023.xlsx')
#workbook.sheetnames
template_sheet = template.active
#building a new BOM



#Create the row filling definition
#this is the bulk of the code functionality, what it does is refrence a template file, and fill in the new file
#it is flexible enough that it allows for multiple sheets to be used in the generated workbook
#the input works for a single row or multiple rows, if using multiple rows the second part of the tuple is NOT included
#e.g. for tuple (i,t) rows i through t-1 are included


def Row_Filler(new_worksheet=['A1'],template_worksheet = ['A1'], rows_numbers = (0,0),relevent_cols = []):
	#the input for rows_numbers is a tuple type, this makes it usable in for loop below if it's an int type
	if type(rows_numbers) == int:
		cell_range1 = rows_numbers
		cell_range2 = rows_numbers + 1
		
	elif type(rows_numbers) == tuple:
		cell_range1 = rows_numbers[0]
		cell_range2 = rows_numbers[1]
	
	for r in range(cell_range1,cell_range2):
		for c in relevent_cols:
		#this copies the values from the template worksheet to the new worksheet
			new_worksheet[f'{c}{r}'] = template_worksheet[f'{c}{r}'].value
		
		#this is all the formatting 
			new_worksheet[f'{c}{r}'].font = copy(template_worksheet[f'{c}{r}'].font)
			new_worksheet[f'{c}{r}'].fill = copy(template_worksheet[f'{c}{r}'].fill)
			new_worksheet[f'{c}{r}'].border = copy(template_worksheet[f'{c}{r}'].border)
			new_worksheet[f'{c}{r}'].alignment = copy(template_worksheet[f'{c}{r}'].alignment)
		#new_worksheet[f'{c}{r}'].alignment =new_worksheet[f'{c}{r}'].alignment.copy(wrapText=True)
			new_worksheet.column_dimensions[f'{c}'].width = copy(template_worksheet.column_dimensions[f'{c}'].width)
		
#building the workbook
bom_workbook = Workbook()
bom_sheet = bom_workbook.active
Row_Filler(bom_sheet,template_sheet,(1,6),relevent_col)
Row_Filler(bom_sheet,template_sheet,(6),relevent_col)
logo = Image("ADRF_Logo.jpg")
bom_sheet.add_image(logo,'A1')
if commercial_or_ps == "Commercial":
	bom_sheet.title = "Commercial BOM"
	bom_sheet['A6'] = project_name + ": Commercial Bill Of Materials"
elif commercial_or_ps == "Public Safety":
	bom_sheet.title = "PS BOM"
	bom_sheet['A6'] = project_name + ": Public Safety Bill Of Materials"

	
def Donor_Equipment():
	#fills in the donor antennas
	if commercial_or_ps == "Commercial":
		Row_Filler(bom_sheet,template_sheet,(9),relevent_col)
		Row_Filler(bom_sheet,template_sheet,(12,14),relevent_col)
	
		#high band then low band antenna
		global donor_high
		donor_high = 0 
		if  'PCS' in bands or 'AWS' in bands or 'BRS' in bands:
			bom_sheet['E12'] = max(pcs_count,aws_count,brs_count)
			donor_high = bom_sheet['E12'].value
		if 'BRS' not in bands and 'PCS' not in bands and 'AWS' not in bands:
			bom_sheet['A12'] = None
			donor_high = 0


		#low band logic 
		global donor_low
		donor_low = 0
		if 'SMR' in bands or '700 MHz' in bands or 'Cellular' in bands: #and '700 MHz' not in bands or 'Cellular' not in bands:
			bom_sheet['E13'] = max(seven_h_mhz_count,cellular_count, smr_count)
			donor_low = bom_sheet['E13'].value
		if '700 MHz' not in bands and 'Cellular' not in bands and 'SMR' not in bands:
			bom_sheet['A13'] = None
			donor_low = 0

		commercial_donor_count = donor_low + donor_high

		if commercial_donor_count != 0:	
			Row_Filler(bom_sheet,template_sheet,(16),relevent_col)			
			bom_sheet['E16'] = commercial_donor_count
			
			Row_Filler(bom_sheet,template_sheet,(18,24),relevent_col)
			for r in range(18,24):
				bom_sheet[f'E{r}'] = commercial_donor_count

			#Donor Coax
			Row_Filler(bom_sheet,template_sheet,(295),relevent_col)
			bom_sheet['E295'] = commercial_donor_count * 150
			#Donor indoor Connector
			Row_Filler(bom_sheet,template_sheet,(300),relevent_col)
			bom_sheet['E300'] = commercial_donor_count
			#Donor Outdoor Connector
			Row_Filler(bom_sheet,template_sheet,(295),relevent_col)
			Row_Filler(bom_sheet,template_sheet,(302),relevent_col)
			bom_sheet['E295'] = commercial_donor_count 
			bom_sheet['E302'] = commercial_donor_count 
			Row_Filler(bom_sheet,template_sheet,(306),relevent_col)
			bom_sheet['E306'] = commercial_donor_count



			if bom_sheet['E12'].value == 0:
				bom_sheet['A12'] = None
			if bom_sheet['E13'].value == 0:
				bom_sheet['A13'] = None
			
			for anc in range(10,23):
				if bom_sheet[f'E{anc}'].value == 0 or bom_sheet[f'E{anc}'].value == None:
					bom_sheet[f'A{anc}'] = None

			bom_sheet['A22'] = None
			bom_sheet['A23'] = None

	if commercial_or_ps == "Public Safety":
		Row_Filler(bom_sheet,template_sheet,(9),relevent_col)
		if '700/800 MHz PS' in ps_bands or 'SMR' in ps_bands or '900MHz with Paging' in ps_bands:
			Row_Filler(bom_sheet,template_sheet,(13),relevent_col)
			bom_sheet['E13'] = 1
		if 'VHF' in ps_bands:
			Row_Filler(bom_sheet,template_sheet,(15),relevent_col)
			if vhf_duplexing == "Yes":
				bom_sheet['E15'] = 1
				Row_Filler(bom_sheet,template_sheet,(83),relevent_col)
				for r in relevent_col:
					bom_sheet[f'{r}83'] = "TBD"
				bom_sheet['E83'] = 2
				bom_sheet['B83'] = "Duplexer"


			elif vhf_duplexing == "No":
				bom_sheet['E15'] = 2


		if 'UHF' in ps_bands:
			Row_Filler(bom_sheet,template_sheet,(14),relevent_col)
			if uhf_duplexing == "Yes":
				bom_sheet['E14'] = 1
				Row_Filler(bom_sheet,template_sheet,(80),relevent_col)
				bom_sheet['E80'] = 2

			elif uhf_duplexing == "No":
				bom_sheet['E14'] = 2


		#ancilliary counter
		donor_antenna_count = 0
		for antennas in range(12,16):
			if bom_sheet[f'E{antennas}'].value != 0 and bom_sheet[f'E{antennas}'].value != None:
				donor_antenna_count += bom_sheet[f'E{antennas}'].value
		Row_Filler(bom_sheet,template_sheet,(17),relevent_col)

		Row_Filler(bom_sheet,template_sheet,(18,24),relevent_col)
		bom_sheet['E17'] = donor_antenna_count
		for donor_anc in range(18,24):
			bom_sheet[f'E{donor_anc}'] = donor_antenna_count
		#Row_Filler(bom_sheet,template_sheet,(295,304),relevent_col)
		#bom_sheet['E295'] = donor_antenna_count
		bom_sheet['E302'] = donor_antenna_count

		#Row_Filler(bom_sheet,template_sheet,(294,304),relevent_col)

		#coax
		Row_Filler(bom_sheet,template_sheet,(294),relevent_col)

		bom_sheet['E294'] = donor_antenna_count * 150
		#Donor indoor Connector
		Row_Filler(bom_sheet,template_sheet,(300),relevent_col)
		bom_sheet['E300'] = donor_antenna_count
		#Donor Outdoor Connector
		
		Row_Filler(bom_sheet,template_sheet,(303),relevent_col)
		Row_Filler(bom_sheet,template_sheet,(302),relevent_col)

		
		bom_sheet['E303'] = donor_antenna_count 
		bom_sheet['E302'] = donor_antenna_count 
		Row_Filler(bom_sheet,template_sheet,(306),relevent_col)
		bom_sheet['E306'] = donor_antenna_count
		#4.3-10 Jumper
		Row_Filler(bom_sheet,template_sheet,(287),relevent_col)
		bom_sheet['E287'] = donor_antenna_count 



		if antenna_mount == "Wall Mount":
			bom_sheet['A20'] = None
			bom_sheet['A21'] = None
		elif antenna_mount == "Non-Pentrating":
			bom_sheet['A22'] = None
			bom_sheet['A23'] = None








#SDR Filling section 
def SDR_24():
		#donor equipment 
		Donor_Equipment()
		#signal source
		#need to add in some if statements to put bands up to user selection
		#SDR NMS
		Row_Filler(bom_sheet,template_sheet,(24,26),relevent_col)
		#SDR700
		if '700 MHz' in bands:
			Row_Filler(bom_sheet,template_sheet,26,relevent_col)
			bom_sheet['E26'] = seven_h_mhz_count
		if 'Cellular' in bands:
			Row_Filler(bom_sheet,template_sheet,27,relevent_col)
			bom_sheet['E27'] = cellular_count
		if 'SMR' in bands:
			Row_Filler(bom_sheet,template_sheet,28,relevent_col)
			bom_sheet['E28'] = smr_count
		if 'PCS' in bands:
			Row_Filler(bom_sheet,template_sheet,29,relevent_col)
			bom_sheet['E29'] = pcs_count
		if 'AWS' in bands:
			Row_Filler(bom_sheet,template_sheet,30,relevent_col)
			bom_sheet['E30'] = aws_count
		if 'BRS' in bands:
			Row_Filler(bom_sheet,template_sheet,(42),relevent_col)
			bom_sheet['E42'] = brs_count
			if project_state == "Canada-West" or  project_state == "Canada-North" or  project_state == "Canada-Central":
				Row_Filler(bom_sheet,template_sheet,(43),relevent_col)
				bom_sheet['E43'] = brs_count
				bom_sheet['A42'] = None 
		if '700MHz' and 'Cellular' in bands or 'PCS' and 'AWS' in bands:
			Row_Filler(bom_sheet,template_sheet,(44),relevent_col)

			bom_sheet['E44'] = (max(seven_h_mhz_count,cellular_count,smr_count,pcs_count,aws_count,brs_count))
			if "T-Mobile" in carriers and 'Verizon' not in carriers and "AT&T" not in carriers:
				bom_sheet['A44'] = None
		if 'AWS' and 'BRS' in tmob_bands:
			Row_Filler(bom_sheet,template_sheet,(46),relevent_col)
			bom_sheet['E46'] = 2
		else:
			bom_sheet['E46'] = 0
				
		#SDR NMS Qty
		#need additional input to make this by carrier for user input 
		bom_sheet['E25'] = math.ceil((seven_h_mhz_count+cellular_count+smr_count+pcs_count+aws_count+brs_count)/4)
		
def SDR_30():
		#donor equipment 
		Donor_Equipment()
		#signal source
		Row_Filler(bom_sheet,template_sheet,(24,26),relevent_col)
		if '700 MHz' in bands:
			Row_Filler(bom_sheet,template_sheet,32,relevent_col)
			bom_sheet['E32'] = seven_h_mhz_count
		if 'Cellular' in bands:
			Row_Filler(bom_sheet,template_sheet,33,relevent_col)
			bom_sheet['E33'] = cellular_count
		if 'SMR' in bands:
			Row_Filler(bom_sheet,template_sheet,34,relevent_col)
			bom_sheet['E34'] = smr_count
		if 'PCS' in bands:
			Row_Filler(bom_sheet,template_sheet,35,relevent_col)
			bom_sheet['E35'] = pcs_count
		if 'AWS' in bands:
			Row_Filler(bom_sheet,template_sheet,36,relevent_col)
			bom_sheet['E36'] = aws_count
		if 'BRS' in bands:
			Row_Filler(bom_sheet,template_sheet,(42),relevent_col)
			bom_sheet['E42'] = brs_count
			if project_state == "Canada-West" or  project_state == "Canada-North" or  project_state == "Canada-Central":
				Row_Filler(bom_sheet,template_sheet,(43),relevent_col)
				bom_sheet['E43'] = brs_count
				bom_sheet['A42'] = None 
		if '700MHz' and 'Cellular' in bands or 'PCS' and 'AWS' in bands:
			Row_Filler(bom_sheet,template_sheet,(44),relevent_col)

			bom_sheet['E44'] = (max(seven_h_mhz_count,cellular_count,smr_count,pcs_count,aws_count,brs_count))
			if "T-Mobile" in carriers and 'Verizon' not in carriers and "AT&T" not in carriers:
				bom_sheet['A44'] = None
		if 'AWS' and 'BRS' in tmob_bands:
			Row_Filler(bom_sheet,template_sheet,(46),relevent_col)
			bom_sheet['E46'] = 2
		else:
			bom_sheet['E46'] = 0
				
		#SDR NMS Qty
		#need additional input to make this by carrier for user input 
		bom_sheet['E25'] = math.ceil((seven_h_mhz_count+cellular_count+smr_count+pcs_count+aws_count+brs_count)/4)

		
def SDR_33():
		#donor equipment 
		Donor_Equipment()
		#signal source
		#signal source
		Row_Filler(bom_sheet,template_sheet,(24,26),relevent_col)
		if '700 MHz' in bands:
			Row_Filler(bom_sheet,template_sheet,37,relevent_col)
			bom_sheet['E37'] = seven_h_mhz_count
		if 'Cellular' in bands:
			Row_Filler(bom_sheet,template_sheet,38,relevent_col)
			bom_sheet['E38'] = cellular_count
		if 'SMR' in bands:
			Row_Filler(bom_sheet,template_sheet,39,relevent_col)
			bom_sheet['E39'] = smr_count
			
		if 'PCS' in bands:
			Row_Filler(bom_sheet,template_sheet,40,relevent_col)
			bom_sheet['E40'] = pcs_count
		if 'AWS' in bands:
			Row_Filler(bom_sheet,template_sheet,41,relevent_col)
			bom_sheet['E41'] = aws_count
		if 'BRS' in bands:
			Row_Filler(bom_sheet,template_sheet,(42),relevent_col)
			bom_sheet['E42'] = brs_count
			if project_state == "Canada-West" or  project_state == "Canada-North" or  project_state == "Canada-Central":
				Row_Filler(bom_sheet,template_sheet,(43),relevent_col)
				bom_sheet['E43'] = brs_count
				bom_sheet['A42'] = None 
		if '700MHz' and 'Cellular' in bands or 'PCS' and 'AWS' in bands:
			Row_Filler(bom_sheet,template_sheet,(44),relevent_col)

			bom_sheet['E44'] = (max(seven_h_mhz_count,cellular_count,smr_count,pcs_count,aws_count,brs_count))
			if "T-Mobile" in carriers and 'Verizon' not in carriers and "AT&T" not in carriers:
				bom_sheet['A44'] = None
		if 'AWS' and 'BRS' in tmob_bands:
			Row_Filler(bom_sheet,template_sheet,(46),relevent_col)
			bom_sheet['E46'] = 2
		else:
			bom_sheet['E46'] = 0
				
		#SDR NMS Qty
		#need additional input to make this by carrier for user input 
		bom_sheet['E25'] = math.ceil((seven_h_mhz_count+cellular_count+smr_count+pcs_count+aws_count+brs_count)/4)
		
def ADXV_HE(square_feet):
	#builds out the HE equipment for ADXV
	Row_Filler(bom_sheet,template_sheet,(95,128),relevent_col)
	# Row_Filler(bom_sheet,template_sheet,(102),relevent_col)
	# Row_Filler(bom_sheet,template_sheet,(105,106),relevent_col)
	# Row_Filler(bom_sheet,template_sheet,(214,217),relevent_col)
	# Row_Filler(bom_sheet,template_sheet,(218),relevent_col)
	# Row_Filler(bom_sheet,template_sheet,(240,242),relevent_col)
	
	#POI Qty
	
	bom_sheet['E105'] = seven_h_l_mhz_count * sectors * mimo_streams
	bom_sheet['E106'] = seven_h_u_mhz_count * sectors * mimo_streams
	bom_sheet['E108'] = smr_count + cellular_count * sectors * mimo_streams
	bom_sheet['E109'] = pcs_count * sectors * mimo_streams
	bom_sheet['E110'] = aws_count * sectors * mimo_streams
	if project_state != "Canada-West" and  project_state != "Canada-North" and  project_state != "Canada-Central":
		bom_sheet['E112'] = brs_count * sectors * mimo_streams
	if project_state == "Canada-West" or  project_state == "Canada-North" or  project_state == "Canada-Central":
		bom_sheet['E115'] = brs_count * sectors * mimo_streams 

	poi_count = 0
	
	#poi count and delete empty rows
	for poi in range(104,119):
		if bom_sheet[f'E{poi}'].value != None:
			poi_count += bom_sheet[f'E{poi}'].value
			if bom_sheet[f'E{poi}'].value == 0:
				bom_sheet[f'A{poi}'] = None

		
	bom_sheet['E126'] = math.ceil((seven_h_mhz_count+cellular_count+smr_count+pcs_count+aws_count+brs_count)/16) * sectors * mimo_streams 
	bom_sheet['E127'] = math.ceil((seven_h_mhz_count+cellular_count+smr_count+pcs_count+aws_count+brs_count)/16) * sectors * mimo_streams


	#counts the total HE modules
	if mpr_or_hpr == "Medium Power Remote":
		odu_count = math.ceil(square_feet/5000/12/4)
		chc_count = (bom_sheet['E126'].value +  bom_sheet['E127'].value) * 2
		nms_count = math.ceil((odu_count + chc_count + poi_count) / 12)
		bom_sheet['E96'] = nms_count * sectors * mimo_streams
		bom_sheet['E102'] = odu_count * sectors * mimo_streams
		
	if mpr_or_hpr == "High Power Remote":
		odu_count = math.ceil(square_feet/5000/24/4)
		chc_count = (bom_sheet['E126'].value +  bom_sheet['E127'].value) * 2
		nms_count = math.ceil((odu_count + chc_count + poi_count) / 12)
		bom_sheet['E96'] = nms_count * sectors * mimo_streams
		bom_sheet['E102'] = odu_count * sectors * mimo_streams
	
	#remove empty rows
	for he_moudle in range(96,128):
		if bom_sheet[f'E{he_moudle}'].value == 0 or bom_sheet[f'E{he_moudle}'].value == None:
				bom_sheet[f'A{he_moudle}'] = None
	
def MPR(square_feet):
	#calculate how many remotes
	remote_count = math.ceil(square_feet/5000/12) * mimo_streams
	if remote_count < sectors * mimo_streams:
		remote_count == sectors * mimo_streams
	if remote_count < num_of_buildings:
		remote_count = num_of_buildings
	#build out the remote
	Row_Filler(bom_sheet,template_sheet,(128,152),relevent_col)
	
	
	#add in the fiber
	Row_Filler(bom_sheet,template_sheet,(196),relevent_col)
	Row_Filler(bom_sheet,template_sheet,(206,212),relevent_col)
	bom_sheet['E206'] = remote_count * 300
	bom_sheet['E207'] = math.ceil(remote_count/6) 
	bom_sheet['E208'] = math.ceil(remote_count/6)*2
	bom_sheet['E209'] = remote_count * 2
	bom_sheet['E210'] = remote_count 
	bom_sheet['E211'] = remote_count * 2

	#add the modules qty
	bom_sheet['E128'] = remote_count
	bom_sheet['E133'] = remote_count
	bom_sheet['E137'] = remote_count
	if '700 MHz' in bands:
		bom_sheet['E141'] = remote_count
	if 'SMR' in bands or 'Cellular' in bands:
		bom_sheet['E143'] = remote_count
	if 'PCS' in bands:
		bom_sheet['E145'] = remote_count
	if 'AWS' in bands:
		bom_sheet['E146'] = remote_count
	if 'BRS' in bands:
		if project_state != "Canada-West" and  project_state != "Canada-North" and  project_state != "Canada-Central":
			bom_sheet['E148'] = remote_count
		if project_state == "Canada-West" or  project_state == "Canada-North" or  project_state == "Canada-Central":
			bom_sheet['E149'] = remote_count

	#CHC-U count 
	if remote_count != 0:
		Row_Filler(bom_sheet,template_sheet,(153),relevent_col)
		bom_sheet['E153'] = remote_count
	if bom_sheet['E46'].value != None:
		bom_sheet['E153'] = remote_count + bom_sheet['E46'].value
		bom_sheet['A46'] = None
	Row_Filler(bom_sheet,template_sheet,(290),relevent_col)
	bom_sheet['E290'] = bom_sheet['E153'].value
	
	#remove empty rows
	for ru_moudle in range(128,154):
		if bom_sheet[f'E{ru_moudle}'].value == 0 or bom_sheet[f'E{ru_moudle}'].value == None:
				bom_sheet[f'A{ru_moudle}'] = None

	
def HPR(square_feet):
	#calculate how many remotes
	remote_count = math.ceil(square_feet/5000/24) * mimo_streams
	if remote_count < sectors * mimo_streams:
		remote_count == sectors * mimo_streams
	if remote_count < num_of_buildings:
		remote_count = num_of_buildings
	#build out the remote
	Row_Filler(bom_sheet,template_sheet,(155,179),relevent_col)

	#add in the fiber
	Row_Filler(bom_sheet,template_sheet,(196),relevent_col)
	Row_Filler(bom_sheet,template_sheet,(206,212),relevent_col)
	bom_sheet['E206'] = remote_count * 300
	bom_sheet['E207'] = math.ceil(remote_count/6) 
	bom_sheet['E208'] = math.ceil(remote_count/6)*2
	bom_sheet['E209'] = remote_count * 2
	bom_sheet['E210'] = remote_count 
	bom_sheet['E211'] = remote_count * 2

	#add the modules qty
	#checks to see if smr and cell are in the unique set "bands" as they are a shared module in adxv, this reduces the bands length by one to check to add a second HPR-CHA
	
	if 'SMR' in bands and 'Cellular' in bands:
		s8c_len = len(bands) - 1
	else:
		s8c_len = len(bands)


	if s8c_len <= 4:
		bom_sheet['E158'] = remote_count
	if s8c_len > 4:
		bom_sheet['E158'] = remote_count * 2

	bom_sheet['E163'] = remote_count
	bom_sheet['E169'] = remote_count

	if '700 MHz' in bands:
		bom_sheet['E172'] = remote_count
	if 'SMR' in bands or 'Cellular' in bands:
		bom_sheet['E173'] = remote_count
	if 'PCS' in bands:
		bom_sheet['E174'] = remote_count
	if 'AWS' in bands:
		bom_sheet['E175'] = remote_count
	if 'BRS' in bands:
		if project_state != "Canada-West" and  project_state != "Canada-North" and  project_state != "Canada-Central":
			bom_sheet['E177'] = remote_count
		
	#CHC-U count 
	if bom_sheet['E46'].value != None:
		bom_sheet['E153'] = remote_count + bom_sheet['E46'].value
		bom_sheet['A46'] = None
	Row_Filler(bom_sheet,template_sheet,(290),relevent_col)
	bom_sheet['E290'] = bom_sheet['E153'].value
	
	#remove empty rows
	for ru_moudle in range(153,195):
		if bom_sheet[f'E{ru_moudle}'].value == 0 or bom_sheet[f'E{ru_moudle}'].value == None:
				bom_sheet[f'A{ru_moudle}'] = None


#Public Safety Definitions
def psr_33():
	Donor_Equipment()
	Row_Filler(bom_sheet,template_sheet,(65),relevent_col)
	Row_Filler(bom_sheet,template_sheet,(69),relevent_col)
	bom_sheet['E69'] = 1
	if backup_time == '12 Hr':
		if wall_mount_battery == False:
			Row_Filler(bom_sheet,template_sheet,(248),relevent_col)
			bom_sheet['E248'] = 1
		if wall_mount_battery == True:
			Row_Filler(bom_sheet,template_sheet,(266),relevent_col)
			bom_sheet['E266'] = 1
	if backup_time == '24 Hr':
		if wall_mount_battery == False:
			Row_Filler(bom_sheet,template_sheet,(248),relevent_col)
			Row_Filler(bom_sheet,template_sheet,(250),relevent_col)
			bom_sheet['E248'] = 2
			bom_sheet['E250'] = 1
		if wall_mount_battery == True:
			Row_Filler(bom_sheet,template_sheet,(267),relevent_col)
			Row_Filler(bom_sheet,template_sheet,(250),relevent_col)
			bom_sheet['E267'] = 2
			bom_sheet['E250'] = 1
	

def psr_37():
	Donor_Equipment()
	Row_Filler(bom_sheet,template_sheet,(65),relevent_col)
	Row_Filler(bom_sheet,template_sheet,(70),relevent_col)
	bom_sheet['E70'] = 1
	if backup_time == '12 Hr':
		if wall_mount_battery == False:
			Row_Filler(bom_sheet,template_sheet,(248),relevent_col)
			bom_sheet['E248'] = 1
		if wall_mount_battery == True:
			Row_Filler(bom_sheet,template_sheet,(266),relevent_col)
			bom_sheet['E266'] = 1
	if backup_time == '24 Hr':
		if wall_mount_battery == False:
			Row_Filler(bom_sheet,template_sheet,(248),relevent_col)
			Row_Filler(bom_sheet,template_sheet,(250),relevent_col)
			bom_sheet['E248'] = 2
			bom_sheet['E250'] = 1
		if wall_mount_battery == True:
			Row_Filler(bom_sheet,template_sheet,(267),relevent_col)
			Row_Filler(bom_sheet,template_sheet,(250),relevent_col)
			bom_sheet['E267'] = 2
			bom_sheet['E250'] = 1

# def ps_smr89(square_feet):
# 	smr_remote_count = 0

# 	Donor_Equipment()
# 	Row_Filler(bom_sheet,template_sheet,(65),relevent_col)
# 	Row_Filler(bom_sheet,template_sheet,(66),relevent_col)
# 	Row_Filler(bom_sheet,template_sheet,(249),relevent_col)
# 	bom_sheet['E66'] = 1
# 	bom_sheet['E249'] = 1
# 	if "900MHz with Paging" in ps_bands:
# 		Row_Filler(bom_sheet,template_sheet,(24),relevent_col)
# 		Row_Filler(bom_sheet,template_sheet,(28),relevent_col)
# 		Row_Filler(bom_sheet,template_sheet,(47),relevent_col)
# 		bom_sheet['E28'] = 1
# 		bom_sheet['E47'] = 1

def fire_u(square_feet):
	ps78_remote_count = 0
	vu_remote_count = 0

	Donor_Equipment()
	Row_Filler(bom_sheet,template_sheet,(65),relevent_col)
	Row_Filler(bom_sheet,template_sheet,(73),relevent_col)
	if "700/800 MHz PS" in ps_bands:
		ps78_remote_count = math.ceil(square_feet/7500/24)
		if ps78_remote_count < num_of_buildings:
			ps78_remote_count = num_of_buildings

		Row_Filler(bom_sheet,template_sheet,(186),relevent_col)
		bom_sheet['E186'] = ps78_remote_count
	
	bom_sheet['E73'] = 1
	
	#battery backup
	if wall_mount_battery == False:
		Row_Filler(bom_sheet,template_sheet,(249),relevent_col)
		Row_Filler(bom_sheet,template_sheet,(248),relevent_col)
		bom_sheet['E248'] = 1
		fire_bbu_count = bom_sheet['E248'].value
	if wall_mount_battery == True:
		Row_Filler(bom_sheet,template_sheet,(266),relevent_col)
		Row_Filler(bom_sheet,template_sheet,(267),relevent_col)
		bom_sheet['E267'] = 1
		fire_bbu_count = bom_sheet['E267'].value
	
	if "VHF" in ps_bands or "UHF" in ps_bands:
		psr_vu()

		vu_bbu_count = 0
		if "700/800 MHz PS" in ps_bands:
			vu_remote_count = ps78_remote_count

			Row_Filler(bom_sheet,template_sheet,(318),relevent_col)
			if vhf_duplexing == "Yes" or uhf_duplexing == "Yes":
				bom_sheet['E318'] = vu_remote_count 
			elif vhf_duplexing == "No" or uhf_duplexing =="No":

				bom_sheet['E318'] = vu_remote_count * 2
				Row_Filler(bom_sheet,template_sheet,(230),relevent_col)
				bom_sheet['E230'] = vu_remote_count
				if wall_mount_battery == False:
					bom_sheet['E249'] = vu_remote_count
				if wall_mount_battery == True:
					bom_sheet['E266'] = vu_remote_count

		else:
			vu_remote_count = math.ceil(square_feet/10000/24)
			if vu_remote_count < num_of_buildings:
				vu_remote_count = num_of_buildings

		
		
		if wall_mount_battery == False:
			vu_bbu_count = bom_sheet['E248'].value
			bom_sheet['E248'].value = fire_bbu_count + vu_bbu_count
		
		#vu remote
		Row_Filler(bom_sheet,template_sheet,(187),relevent_col)
		bom_sheet['E187'] = vu_remote_count
		
		

	ps_remote_count = ps78_remote_count + vu_remote_count

	#fiber add in
	Row_Filler(bom_sheet,template_sheet,(196),relevent_col)
	Row_Filler(bom_sheet,template_sheet,(206,212),relevent_col)
	bom_sheet['E206'] = ps_remote_count * 300
	bom_sheet['E207'] = math.ceil(ps_remote_count/6) 
	bom_sheet['E208'] = math.ceil(ps_remote_count/6)+ps_remote_count
	bom_sheet['E209'] = ps_remote_count * 2
	bom_sheet['E210'] = ps_remote_count 
	bom_sheet['E211'] = ps_remote_count * 2
	# jumpers for remotes
	bom_sheet['E287'] = bom_sheet['E287'].value + ps_remote_count



		







	#OEU logic
	if ps_remote_count > 8:
		Row_Filler(bom_sheet,template_sheet,(74),relevent_col)
		Row_Filler(bom_sheet,template_sheet,(183),relevent_col)
		bom_sheet['E74'] = math.ceil((ps_remote_count-8)/8) 
		bom_sheet['E183'] = bom_sheet['E73'].value

	if wall_mount_battery == False:
		bom_sheet['E249'] = ps_remote_count
	if wall_mount_battery == True:
		bom_sheet['E266'] = ps_remote_count




def psr_vu():
	Donor_Equipment()
	
	Row_Filler(bom_sheet,template_sheet,(65),relevent_col)
	
	if "UHF" in ps_bands or "VHF" in ps_bands:
		#battery backup
		if wall_mount_battery == False:
			Row_Filler(bom_sheet,template_sheet,(248),relevent_col)
			Row_Filler(bom_sheet,template_sheet,(251),relevent_col)
		if wall_mount_battery == True:
			Row_Filler(bom_sheet,template_sheet,(269),relevent_col)
			Row_Filler(bom_sheet,template_sheet,(250),relevent_col)

		if "700/800 MHz PS" in ps_bands:
			Row_Filler(bom_sheet,template_sheet,(318),relevent_col)
			bom_sheet['E318'] = 1
			if vhf_duplexing == "No" or uhf_duplexing =="No":
				bom_sheet['E318'] = 2


	if "VHF" in ps_bands:
		
		#VHF repeater
		Row_Filler(bom_sheet,template_sheet,(76),relevent_col)
		bom_sheet['E76'] = 1

		#battery backup		
		if wall_mount_battery == False:
			bom_sheet['E248'] = 2
			bom_sheet['E251'] = 1
		if wall_mount_battery == True:
			if backup_time == '12 Hr':
				bom_sheet['E269'] = 1
			if backup_time == '24 Hr':
				bom_sheet['E269'] = 2
				bom_sheet['E250'] = 1

	if "UHF" in ps_bands:

		if uhf_duplexing == "No":
			Row_Filler(bom_sheet,template_sheet,(76),relevent_col)
			bom_sheet['E76'] = 1

			

		if uhf_duplexing == "Yes":
			Row_Filler(bom_sheet,template_sheet,(77),relevent_col)
			bom_sheet['E77'] = 1

		if wall_mount_battery == False:
			bom_sheet['E248'] = 2
			bom_sheet['E251'] = 1
		if wall_mount_battery == True:
			if backup_time == '12 Hr':
				bom_sheet['E269'] = 1
			if backup_time == '24 Hr':
				bom_sheet['E250'] = 1
				bom_sheet['E269'] = 2
	
	if "UHF" in ps_bands and "VHF" in ps_bands:
		#vhf/uhf diplexer
		Row_Filler(bom_sheet,template_sheet,(317),relevent_col)
		
		if uhf_duplexing == "No":
			Row_Filler(bom_sheet,template_sheet,(76),relevent_col)
			bom_sheet['E76'] = 2

			

		#battery backup
		if wall_mount_battery == False:
			bom_sheet['E248'] = 4
			bom_sheet['E251'] = 2
		if wall_mount_battery == True:
			if backup_time == '12 Hr':
				bom_sheet['E269'] = 2
			if backup_time == '24 Hr':
				bom_sheet['E250'] = 2
				bom_sheet['E269'] = 4


		#diplexer logic count
		if vhf_duplexing == "Yes" or uhf_duplexing == "Yes":
			bom_sheet['E317'] = 1
			if "700/800 MHz PS" in ps_bands:
				Row_Filler(bom_sheet,template_sheet,(318),relevent_col)
				bom_sheet['E318'] = 1


		if vhf_duplexing == "No" or uhf_duplexing =="No":
			bom_sheet['E31'] = 2
			if "700/800 MHz PS" in ps_bands:
				#this value gets overwritten when the system switches to active in the fireu function
				Row_Filler(bom_sheet,template_sheet,(318),relevent_col)
				bom_sheet['E318'] = 2






		

def Passives_Count(square_feet):

	#the logic for all passives is that the total count of passives is the number of antennas - amplifiers, for passives systems this is antennas - 1, since there is only (1) amplifier
	#the first 3 passive devices are 2-way splitters, couplers are not added until total antennas > (4), as 3 2ws yield 4 antennas
	#public safety and commercial are split to allow them to have different sets of passive
	#nested in the ps is vhf/uhf which swap the 698MHz-2700MHz to the microlab


	if commercial_or_ps == "Commercial":
		Row_Filler(bom_sheet,template_sheet,(225),relevent_col)

		ant_count = math.ceil(square_feet/5000)
		bom_sheet['E227'] = 0
		#donor antenna check
		if donor_low != 0 and donor_high != 0:
			Row_Filler(bom_sheet,template_sheet,(227),relevent_col)
			bom_sheet['E227'] = min(donor_low, donor_high)
		elif donor_low == 0 and  donor_high == 0:
			bom_sheet['A227'] = None


		Row_Filler(bom_sheet,template_sheet,(237),relevent_col)
		if ant_count < num_of_floors:
			ant_count = num_of_floors
		if ant_count < 2:
			bom_sheet['A237'] = None
		else: 
			bom_sheet['E237'] = math.ceil(ant_count-1)
		bom_sheet['E240'] = 0
		bom_sheet['E241'] = 0

		if square_feet <= 100000:
			if ant_count > 4 and ant_count <= 12:
				bom_sheet['E237'] = 3
				Row_Filler(bom_sheet,template_sheet,(240),relevent_col)
				bom_sheet['E240'] = ant_count-1-bom_sheet['E241'].value
			if ant_count > 12 and ant_count <= 20:
				bom_sheet['E241'] = 3
				Row_Filler(bom_sheet,template_sheet,(240),relevent_col)
				Row_Filler(bom_sheet,template_sheet,(234),relevent_col)
				bom_sheet['E241'] = (ant_count-8)-1-bom_sheet['E241'].value
				bom_sheet['E240'] = ant_count-1-bom_sheet['E241'].value-bom_sheet['E241'].value
			

		
		elif square_feet > 100000 or num_of_buildings > 1:
			if mpr_or_hpr == "Medium Power Remote":
				remote_count = math.ceil(square_feet/5000/12)
				ant_count = math.ceil(square_feet/5000)
				ant_count_remote = math.ceil(ant_count/remote_count)

				if ant_count < 2: 
					bom_sheet['A237'] = None
				else: 
					bom_sheet['E237'] = math.ceil(ant_count_remote-1)
				
				if ant_count_remote > 4 and ant_count_remote <= 12:
					bom_sheet['E237'] = 3
					Row_Filler(bom_sheet,template_sheet,(240),relevent_col)
					bom_sheet['E240'] = ant_count_remote-1-bom_sheet['E237'].value
				if ant_count_remote > 12 and ant_count_remote <= 20:
					bom_sheet['E237'] = 3
					Row_Filler(bom_sheet,template_sheet,(240),relevent_col)
					Row_Filler(bom_sheet,template_sheet,(241),relevent_col)
					bom_sheet['E241'] = (ant_count_remote-8)-1-bom_sheet['E237'].value
					bom_sheet['E240'] = ant_count_remote-1-bom_sheet['E237'].value-bom_sheet['E241'].value
			
			if mpr_or_hpr == "High Power Remote":
				remote_count = math.ceil(square_feet/5000/24)
				ant_count = math.ceil(square_feet/5000)
				ant_count_remote = math.ceil(ant_count/remote_count)

				if ant_count < 2: 
					bom_sheet['A237'] = None
				else: 
					bom_sheet['E237'] = math.ceil(ant_count_remote-1)
				
				if ant_count_remote > 4 and ant_count_remote <= 12:
					bom_sheet['E237'] = 3
					Row_Filler(bom_sheet,template_sheet,(240),relevent_col)
					bom_sheet['E240'] = ant_count_remote-1-bom_sheet['E237'].value
				if ant_count_remote > 12 and ant_count_remote <= 20:
					bom_sheet['E237'] = 3
					Row_Filler(bom_sheet,template_sheet,(240),relevent_col)
					Row_Filler(bom_sheet,template_sheet,(241),relevent_col)
					bom_sheet['E241'] = (ant_count_remote-8)-1-bom_sheet['E237'].value
					bom_sheet['E240'] = ant_count_remote-1-bom_sheet['E237'].value-bom_sheet['E241'].value

		Row_Filler(bom_sheet,template_sheet,(243),relevent_col)
		
		if ant_count < num_of_floors:
				ant_count = num_of_floors
		bom_sheet['E243'] = ant_count

		


		if square_feet <= 100000 and num_of_buildings < 2:
			if bom_sheet['E44'].value != None:
				if bom_sheet['E44'].value <= 1:
					pass
				if bom_sheet['E44'].value > 1 and bom_sheet['E44'].value <= 2:
					Row_Filler(bom_sheet,template_sheet,(242),relevent_col)
					bom_sheet['E242'].value = 1
				if bom_sheet['E44'].value > 2:
					Row_Filler(bom_sheet,template_sheet,(314),relevent_col)
					bom_sheet['E314'].value = 1
			

		
		Row_Filler(bom_sheet,template_sheet,(292),relevent_col)
		
		if bom_sheet['E292'].value == None:
			bom_sheet['E292'].value =0
			bom_sheet['A292'].value = None
		if bom_sheet['E314'].value == None:
			bom_sheet['E314'].value =0
			bom_sheet['A314'].value = None
		
		bom_sheet['E292'] = math.ceil(ant_count*.75) + bom_sheet['E314'].value
		Row_Filler(bom_sheet,template_sheet,(295),relevent_col)
		bom_sheet['E295'] = ant_count*150
		Row_Filler(bom_sheet,template_sheet,(306),relevent_col)
		if ant_count>1:
			bom_sheet['E306'] = round(ant_count*.25) + round((bom_sheet['E237'].value +  bom_sheet['E227'].value + bom_sheet['E240'].value + bom_sheet['E241'].value) * 3)
		elif ant_count < 2:
			bom_sheet['E306'] = math.ceil(ant_count*.25) +  (bom_sheet['E227'].value + bom_sheet['E240'].value + bom_sheet['E241'].value) * 3

		if bom_sheet['E306'].value == None or bom_sheet['E306'].value ==0:
			#bom_sheet['E403'].value = 0
			bom_sheet['A306'].value = None
		if siso_or_mimo == "MIMO":
			#st.write("MIMO check",bom_sheet['E237'].value * mimo_streams )
			for passives_amount in range(237,243):
				if bom_sheet[f'E{passives_amount}'].value == None:
					bom_sheet[f'E{passives_amount}'] = 0
				bom_sheet[f'E{passives_amount}'] = bom_sheet[f'E{passives_amount}'].value * mimo_streams 
			for passives_amount in range(287,308):
				if bom_sheet[f'E{passives_amount}'].value == None:
					bom_sheet[f'E{passives_amount}'] = 0
				bom_sheet[f'E{passives_amount}'] = bom_sheet[f'E{passives_amount}'].value * mimo_streams 
			bom_sheet['A243'] = None
			Row_Filler(bom_sheet,template_sheet,(332),relevent_col)
			bom_sheet['E332'] = round(bom_sheet['E243'].value)
		

	if commercial_or_ps == "Public Safety":
		Row_Filler(bom_sheet,template_sheet,(225),relevent_col)
		if "VHF" not in ps_bands and "UHF" not in ps_bands:
			ps_remote_count = math.ceil(square_feet/7500/24)
			
			#this sets the remote count to match the buildings if there are more than one
			if ps_remote_count < num_of_buildings:
				ps_remote_count = num_of_buildings

			ps_antenna_count = math.ceil(square_feet/7500)
			Row_Filler(bom_sheet,template_sheet,(236),relevent_col)
			
			if ps_antenna_count < num_of_floors:
				ps_antenna_count = num_of_floors
			bom_sheet['E236'] = ps_antenna_count

			
			bom_sheet['E233'] = 0
			bom_sheet['E234'] = 0
			if square_feet <= 375000:
				#this accounts for BDA fed systems, if it's active the remote count takes over
				if num_of_buildings <=1:
					ps_remote_count = 1
				if ps_antenna_count < 2:
					bom_sheet['E234'] = 0
				if ps_antenna_count >= 2:
					Row_Filler(bom_sheet,template_sheet,(230),relevent_col)
					bom_sheet['E230'] = math.ceil(ps_antenna_count-ps_remote_count)
					
				if ps_antenna_count > 4 and ps_antenna_count <= 12:
					bom_sheet['E230'] = 3
					Row_Filler(bom_sheet,template_sheet,(233),relevent_col)
					bom_sheet['E233'] = ps_antenna_count-ps_remote_count-bom_sheet['E230'].value
				if ps_antenna_count > 12 and ps_antenna_count <= 20:
					bom_sheet['E230'] = ps_antenna_count - 10
					Row_Filler(bom_sheet,template_sheet,(233),relevent_col)
					Row_Filler(bom_sheet,template_sheet,(234),relevent_col)
					bom_sheet['E234'] = (ps_antenna_count-8)-ps_remote_count-bom_sheet['E230'].value
					bom_sheet['E233'] = ps_antenna_count-ps_remote_count-bom_sheet['E230'].value-bom_sheet['E234'].value


				if ps_antenna_count > 20:
					bom_sheet['E230'] = ps_antenna_count - ps_remote_count - math.ceil(ps_antenna_count*.70)
					Row_Filler(bom_sheet,template_sheet,(233),relevent_col)
					Row_Filler(bom_sheet,template_sheet,(234),relevent_col)
					bom_sheet['E234'] = math.ceil(ps_antenna_count)-ps_remote_count-math.ceil(ps_antenna_count*.30)-bom_sheet['E230'].value
					bom_sheet['E233'] = math.ceil(ps_antenna_count)-ps_remote_count-bom_sheet['E230'].value-bom_sheet['E234'].value
				
				#checks for negatives and gives accurate count, if any of the couplers/splitters are negative they get added to the next value until they are zeroed out
				if bom_sheet['E230'].value == None:
					bom_sheet['E230'].value = 0

				if bom_sheet['E234'].value <= 0:
					bom_sheet['E233'] = bom_sheet['E233'].value + bom_sheet['E234'].value
					bom_sheet['A234'] = None
				if bom_sheet['E233'].value <= 0:
					bom_sheet['E230'] = bom_sheet['E233'].value + bom_sheet['E230'].value
					bom_sheet['A233'] = None
				if bom_sheet['E230'].value <= 0:
					bom_sheet['A230'] = None

			if square_feet > 375000:
				Row_Filler(bom_sheet,template_sheet,(230),relevent_col)
				Row_Filler(bom_sheet,template_sheet,(233),relevent_col)
				Row_Filler(bom_sheet,template_sheet,(234),relevent_col)
				bom_sheet['E230'] = ps_antenna_count - ps_remote_count - math.ceil(ps_antenna_count*.70)
				bom_sheet['E234'] = math.ceil(ps_antenna_count)-ps_remote_count-math.ceil(ps_antenna_count*.30)-bom_sheet['E230'].value
				bom_sheet['E233'] = math.ceil(ps_antenna_count)-ps_remote_count-bom_sheet['E230'].value-bom_sheet['E234'].value

			Row_Filler(bom_sheet,template_sheet,(291),relevent_col)
			Row_Filler(bom_sheet,template_sheet,(296),relevent_col)
			bom_sheet['E291'] = math.ceil(ps_antenna_count*.75)
			
			bom_sheet['E296'] = ps_antenna_count*150
			
			Row_Filler(bom_sheet,template_sheet,(300),relevent_col)
			if ps_antenna_count>1:
				bom_sheet['E300'] = round(ps_antenna_count*.25) + round((bom_sheet['E230'].value +  bom_sheet['E233'].value + bom_sheet['E234'].value) * 3)
			elif ps_antenna_count < 2:
				bom_sheet['E300'] = math.ceil(ps_antenna_count*.25) +  (bom_sheet['E233'].value + bom_sheet['E234'].value) * 3
		
		if "VHF" in ps_bands or "UHF" in ps_bands and '700/800 MHz PS' not in ps_bands and  'SMR' not in ps_bands and '900MHz with Paging' not in ps_bands:
			bom_sheet['E320'] = 0
			bom_sheet['E325'] = 0
			bom_sheet['E324'] = 0
			ps_remote_count = math.ceil(square_feet/10000/24)
			if ps_remote_count < num_of_buildings:
				ps_remote_count = num_of_buildings
			if vhf_duplexing == "No" and uhf_duplexing == "No":
				vhf_uhf_antenna_count = math.ceil(square_feet/10000)
				if vhf_uhf_antenna_count < 2:
					vhf_uhf_antenna_count = 2


			if vhf_duplexing == "Yes" or uhf_duplexing == "Yes":
				vhf_uhf_antenna_count = math.ceil(square_feet/15000)
			if vhf_uhf_antenna_count < num_of_floors:
				vhf_uhf_antenna_count = num_of_floors
				

			if square_feet <= 450000:

				if vhf_uhf_antenna_count < 2:
					bom_sheet['E325'] = 0
				if vhf_uhf_antenna_count >= 2:
					Row_Filler(bom_sheet,template_sheet,(320),relevent_col)
					bom_sheet['E320'] = math.ceil(vhf_uhf_antenna_count-ps_remote_count)
							
				if vhf_uhf_antenna_count > 4 and vhf_uhf_antenna_count <= 12:
					bom_sheet['E320'] = 3
					Row_Filler(bom_sheet,template_sheet,(324),relevent_col)
					bom_sheet['E324'] = vhf_uhf_antenna_count-ps_remote_count-bom_sheet['E320'].value
					if vhf_uhf_antenna_count > 12 and vhf_uhf_antenna_count <= 20:
						bom_sheet['E320'] = vhf_uhf_antenna_count - 10
						Row_Filler(bom_sheet,template_sheet,(324),relevent_col)
						Row_Filler(bom_sheet,template_sheet,(325),relevent_col)
						#10 dB coupler
						bom_sheet['E325'] = (vhf_uhf_antenna_count-8)-ps_remote_count-bom_sheet['E320'].value
						#6 dB coupler
						bom_sheet['E324'] = vhf_uhf_antenna_count-ps_remote_count-bom_sheet['E324'].value-bom_sheet['E320'].value
				if vhf_uhf_antenna_count > 20:
					bom_sheet['E320'] = vhf_uhf_antenna_count - ps_remote_count - math.ceil(vhf_uhf_antenna_count*.70)
					Row_Filler(bom_sheet,template_sheet,(325),relevent_col)
					Row_Filler(bom_sheet,template_sheet,(324),relevent_col)
					bom_sheet['E325'] = math.ceil(vhf_uhf_antenna_count)-ps_remote_count-math.ceil(vhf_uhf_antenna_count*.30)-bom_sheet['E320'].value
					bom_sheet['E324'] = math.ceil(vhf_uhf_antenna_count)-ps_remote_count-bom_sheet['E320'].value-bom_sheet['E325'].value
				

				

				
			if square_feet > 450000:
				
				Row_Filler(bom_sheet,template_sheet,(320),relevent_col)
				Row_Filler(bom_sheet,template_sheet,(324),relevent_col)
				Row_Filler(bom_sheet,template_sheet,(325),relevent_col)
				bom_sheet['E320'] = vhf_uhf_antenna_count - ps_remote_count - math.ceil(vhf_uhf_antenna_count*.70)
				bom_sheet['E325'] = math.ceil(vhf_uhf_antenna_count)-ps_remote_count-math.ceil(vhf_uhf_antenna_count*.30)-bom_sheet['E320'].value
				bom_sheet['E324'] = math.ceil(vhf_uhf_antenna_count)-ps_remote_count-bom_sheet['E320'].value-bom_sheet['E325'].value

			#Coax/		
			Row_Filler(bom_sheet,template_sheet,(291),relevent_col)
			Row_Filler(bom_sheet,template_sheet,(296),relevent_col)
			bom_sheet['E291'] = math.ceil(vhf_uhf_antenna_count*.75)
			bom_sheet['E296'] = vhf_uhf_antenna_count*150

			#jumpers
			if bom_sheet['E287'].value == None:
				Row_Filler(bom_sheet,template_sheet,(287),relevent_col)
				bom_sheet['E287'] = ps_remote_count
			if bom_sheet['E287'].value > 0:
				bom_sheet['E287'] = bom_sheet['E287'].value + ps_remote_count


			#connectors
			Row_Filler(bom_sheet,template_sheet,(300),relevent_col)
			if vhf_uhf_antenna_count>1:
				bom_sheet['E300'] = round(vhf_uhf_antenna_count*.25) + round((bom_sheet['E320'].value +  bom_sheet['E325'].value + bom_sheet['E324'].value) * 3)
			elif vhf_uhf_antenna_count < 2:
				bom_sheet['E300'] = math.ceil(vhf_uhf_antenna_count*.25) +  (bom_sheet['E325'].value + bom_sheet['E324'].value) * 3

			#antennas
			Row_Filler(bom_sheet,template_sheet,(330),relevent_col)
			bom_sheet['E330'] = vhf_uhf_antenna_count
			#checks for negatives and gives accurate count, if any of the couplers/splitters are negative they get added to the next value until they are zeroed out
			if bom_sheet['E325'].value < 0:
				bom_sheet['E324'] = bom_sheet['E324'].value + bom_sheet['E325'].value
				bom_sheet['A325'] = None
			if bom_sheet['E324'].value < 0:
				bom_sheet['E320'] = bom_sheet['E324'].value + bom_sheet['E320'].value
				bom_sheet['A324'] = None
			if bom_sheet['E320'].value < 0:
				bom_sheet['A320'] = None







		
		
# set the headers
#these are the coloumn that contain the information from the template BOM
relevent_col=['A','B','C','D','E']


Row_Filler(bom_sheet,template_sheet,(8),relevent_col)
#for c in relevent_col:
	#bom_sheet[f'{c}8']=template_sheet[f'{c}8'].value

#commercial BOM building
if commercial_or_ps == "Commercial":
	if num_of_buildings < 2:
		if square_feet <= 50000:
			
			SDR_24()	
			#passives
			Passives_Count(square_feet)
			
		if square_feet > 50000 and square_feet <= 75000:
			SDR_30()
			#passives
			Passives_Count(square_feet)
			
		if square_feet > 75000 and square_feet <= 100000:
			SDR_33()
			#passives
			Passives_Count(square_feet)
	if square_feet > 100000 or num_of_buildings > 1:
		if signal_source == "Repeaters":
			SDR_24()
		if mpr_or_hpr == "Medium Power Remote":
			ADXV_HE(square_feet)
			MPR(square_feet)
		 
		if mpr_or_hpr == "High Power Remote":
			ADXV_HE(square_feet)
			HPR(square_feet)
		
		#passives
		Passives_Count(square_feet)
if commercial_or_ps == "Public Safety":
	if num_of_buildings < 2:
		if "700/800 MHz PS" in ps_bands:
			if square_feet <= 225000:
				psr_33()
				Passives_Count(square_feet)
			if square_feet > 225000 and square_feet <= 375000:
				psr_37()
				Passives_Count(square_feet)
	


	if square_feet > 375000 and square_feet <= 6480000 or  num_of_buildings > 1:
		fire_u(square_feet)
		Passives_Count(square_feet)
		if square_feet > 6480000:
				st.write("Project Too Large, reach out to DASET@adrftech.com")
	
	if "VHF" or "UHF" in ps_bands:
		if num_of_buildings < 2:
			if square_feet <= 450000:
				psr_vu()
				Passives_Count(square_feet)

		if square_feet > 450000 and square_feet <= 8640000 or  num_of_buildings > 1:
			fire_u(square_feet)
			Passives_Count(square_feet)

			if "700/800 MHz PS" in ps_bands:
				if square_feet > 3240000:
					st.write("Project Too Large, reach out to DASET@adrftech.com")
		elif square_feet > 8640000:
			st.write("Project Too Large, reach out to DASET@adrftech.com")
	if epo_switch == True:
		Row_Filler(bom_sheet,template_sheet,(84),relevent_col)
		Row_Filler(bom_sheet,template_sheet,(85),relevent_col)
		psr_count = 0
		for psrs in range(69,78):
			if bom_sheet[f'E{psrs}'].value == None:
				bom_sheet[f'E{psrs}'] = 0
			psr_count += bom_sheet[f'E{psrs}'].value

		bom_sheet['E84'] = psr_count
		bom_sheet['E85'] = bom_sheet['E84'].value



 #last row
Row_Filler(bom_sheet,template_sheet,(341),relevent_col)
Row_Filler(bom_sheet,template_sheet,(355,357),relevent_col)
bom_sheet['A341'] = ' '
bom_sheet['A355'] = ' '
bom_sheet['A356'] = ' '	
bom_sheet['A7'] = ' '


#delete rows

index_row = []

# loop each row in column A
for i in range(6, bom_sheet.max_row):
	# define emptiness of cell
	if bom_sheet.cell(i, 1).value is None:
		# collect indexes of rows
		index_row.append(i)

# loop each index value
for row_del in range(len(index_row)):
	bom_sheet.delete_rows(idx=index_row[row_del], amount=1)
	# exclude offset of rows through each iteration
	index_row = list(map(lambda k: k - 1, index_row))

for row in range(1,bom_sheet.max_row+1):
		bom_sheet.row_dimensions[row].height = 25 #copy(template_worksheet.row_dimensions[r].height)

#makes the last row bigger
last_row = bom_sheet.max_row
bom_sheet.row_dimensions[last_row].height = 80


#date = datetime.datetime.now()
date = datetime.datetime.now(pytz.timezone('US/Central'))
#date=date.astimezone(pytz.timezone('US/Central')).strftime('%Y%mm%dd')
#save the file
today = date.strftime("%Y")+"/"+date.strftime("%m")+"/"+str(date.strftime("%d"))
new_bom_workbook = "ADRF_"+project_name+"_ROM BOM_"+date.strftime("%Y")+date.strftime("%m")+str(date.strftime("%d"))+".xlsx"
bom_workbook.save(filename=new_bom_workbook)


with st.container():
	st.write("---")
	st.subheader('ROM BOM:')
	read_new_bom_workbook = pd.read_excel(new_bom_workbook)
	st.write(read_new_bom_workbook)

with NamedTemporaryFile(delete=False) as tmp:
	bom_workbook.save(tmp.name)
	data = BytesIO(tmp.read())

#google sheets API
scope = ["https://spreadsheets.google.com/feeds",'https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"]

creds = ServiceAccountCredentials.from_json_keyfile_name("creds.json", scope)

client = gspread.authorize(creds)

gsheet = client.open("BOM Tool Contact Information").sheet1  # Open the spreadhseet

if commercial_or_ps == "Public Safety":
	freqs = ps_bands
	carriers = ["Public Safety AHJ"]
if commercial_or_ps == "Commercial":
	freqs = bands


def project_tracker(company,name,email,phone):
	
	data = gsheet.get_all_records()
	new_row = [str(company),str(name),str(email),str(phone),str(project_name),str(project_state),str(commercial_or_ps),str(square_feet),str(num_of_buildings), ', '.join(carriers),', '.join(freqs),str(today)]
	
	#new_row = []
	#end_row = len(data)+1
	#st.write(str(new_row)+str(end_row))
	gsheet.append_row(new_row,table_range="A2")
	send_to_email_addresses.append(email)
	(send_email(user_name,send_to_email_addresses,user_company))




# #Email Sending

def send_email(name,send_to,company):
	message = f"""
{name},

See attached ROM for {project_name}.

If you have any questions please reach out to daset@adrftech.com\n
Your ADRF representative can be reached at {rsm_email}\n

Project Information:\n
State: {project_state}\n
Region: {project_region}\n
Type: {commercial_or_ps}\n
Carriers: {carriers}\n
Square Feet: {square_feet}\n
Number of Buildings: {num_of_buildings}\n
Number of Floors: {num_of_floors}\n
Frequencies: {freqs}\n

	
DAS Engineering Team
Advanced RF Technologies, Inc.
3116 W. Vanowen St.
Burbank, CA 91505

Office: (818) 840-8131
24/7: (800) 320-9345

The information contained in this message may be privileged and confidential and protected from disclosure. If the reader of this message is not the intended recipient, or an employee or agent responsible for delivering this message to the intended recipient, you are hereby notified that any dissemination, distribution or copying of this communication is strictly prohibited. If you have received this e-mail in error, please notify us immediately by replying to the message and deleting it from your computer. Thank you. Advanced RF Technologies, Inc.
	

	"""
	files=new_bom_workbook
	username = daset_email
	password = daset_password
	
	subject = f'ADRF ROM BOM for {project_name} with {company}'
	
	msg = EmailMessage()
	msg['Subject'] = subject
	msg['From'] = username
	msg['To'] = send_to #', '.join(send_to)
	msg.set_content(message)

	with open(files,'rb') as f:
		file_data = f.read()
		file_type = 'xlsx'

	msg.add_attachment(file_data,maintype="application",subtype=file_type,filename=new_bom_workbook)

	with smtplib.SMTP_SSL('smtp.gmail.com',465) as smtp:
		smtp.login(username, password)
		smtp.send_message(msg)


#------------contact form
with st.container():
	
	st.write("---")
	st.header("Contact Information")
	st.write("##")
	
	left_column, left_middle_column = st.columns(2)
	with left_column:
		user_name = st.text_input("Name")
		user_company = st.text_input("Company")
	with left_middle_column:
		user_phone = st.text_input("Phone Number")
		user_email = st.text_input("Email")
	if "@" in user_email: 
		email_check = False
	if "@" not in user_email:
		email_check = True
	sheet_write = st.button("Click to Submit Information and Download BOM",
		disabled=email_check)	
	if sheet_write == True:
		email_send = st.download_button("Download ROM BOM",
			data=data,
			mime='xlsx',
			file_name=new_bom_workbook,
			on_click=(project_tracker(user_company,user_name,user_email,user_phone)),
			disabled=email_check
			)
		
st.write("####")
st.write('Please contact daset@adrftech.com for any inquiries.')
st.write('---')
st.write("####")

os.remove(new_bom_workbook)
